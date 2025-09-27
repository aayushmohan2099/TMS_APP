# bmmu/management/commands/check_trainingplan_conflicts.py
import csv
from pathlib import Path
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from bmmu.models import TrainingPlan

EXPECTED_HEADERS = {
    "training_name": ["training_name", "module", "module(training name)", "module (training name)"]
}

def header_lookup_map(fieldnames):
    lower_headers = {h.strip().lower(): h for h in fieldnames if h}
    mapping = {}
    for key, poss in EXPECTED_HEADERS.items():
        found = None
        for cand in poss:
            if cand.strip().lower() in lower_headers:
                found = lower_headers[cand.strip().lower()]
                break
        mapping[key] = found
    return mapping

class Command(BaseCommand):
    help = "List TrainingPlan names in input file that already exist in DB (match by training_name)."

    def add_arguments(self, parser):
        parser.add_argument("file", help="Path to CSV or XLSX file")
        parser.add_argument("--format", choices=("csv", "xlsx"), default=None)

    def handle(self, *args, **opts):
        p = Path(opts["file"])
        if not p.exists():
            self.stderr.write("File not found.")
            return 1
        fmt = opts["format"]
        if fmt is None:
            if p.suffix.lower() == ".csv":
                fmt = "csv"
            elif p.suffix.lower() in (".xls", ".xlsx"):
                fmt = "xlsx"
            else:
                self.stderr.write("Use --format csv|xlsx")
                return 1

        names = []
        if fmt == "csv":
            with open(p, newline='', encoding='utf-8') as fh:
                reader = csv.DictReader(fh)
                map_hdr = header_lookup_map(reader.fieldnames)
                for row in reader:
                    hdr = map_hdr.get("training_name")
                    if hdr:
                        names.append(str(row.get(hdr) or "").strip())
        else:
            wb = load_workbook(filename=str(p), read_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                self.stderr.write("Empty xlsx")
                return 1
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            map_hdr = header_lookup_map(headers)
            idx = headers.index(map_hdr.get("training_name")) if map_hdr.get("training_name") in headers else None
            for row in it:
                if idx is not None:
                    names.append(str(row[idx] or "").strip())

        names = [n for n in names if n]
        existing = set(TrainingPlan.objects.filter(training_name__in=names).values_list('training_name', flat=True))
        self.stdout.write(f"Total rows checked: {len(names)}")
        self.stdout.write(f"Existing in DB (count={len(existing)}):")
        for n in sorted(existing):
            self.stdout.write(" - " + n)
        missing = [n for n in names if n not in existing]
        self.stdout.write(f"\nNot present in DB (count={len(missing)}): {len(missing)}")
        return 0
