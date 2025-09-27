# bmmu/management/commands/import_trainingplans.py
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from django.core.management.base import BaseCommand
from django.db import IntegrityError

from bmmu.models import TrainingPlan

# --- Helpers: normalization maps and parsing ----------------

_TYPE_MAP = {
    "res": "RES",
    "residential": "RES",
    "nonres": "NON RES",
    "nonresidential": "NON RES",
    "nonres": "NON RES",
    "nonresidential": "NON RES",
    "nonresidential": "NON RES",
    "nonresidential": "NON RES",
    "nonres": "NON RES",
    "nonresidential": "NON RES",
    "nonres": "NON RES",
    "non_res": "NON RES",
    "nonresidential": "NON RES",
    "other": "OTHER",
}

_LEVEL_RULES = [
    ("VILLAGE", ["village"]),
    ("SHG", ["shg"]),
    ("CLF", ["clf"]),
    ("BLOCK_DISTRICT", ["block", "district"]),
    ("CMTC/BLOCK", ["cmtc"]),
    ("BLOCK", ["block"]),
    ("DISTRICT", ["district"]),
    ("STATE", ["state"]),
    ("WITHIN_STATE", ["within state", "within_state", "within"]),
    ("OUTSIDE_STATE", ["outside state", "outside_state", "outside"]),
]

_APPROVAL_MAP = {
    "sanctioned": "SANCTIONED",
    "pending": "PENDING",
    "denied": "DENIED",
    "reject": "DENIED",
}

DATE_FORMATS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y", "%d %b %Y", "%d %B %Y")

def _clean_key(s):
    if s is None:
        return "", ""
    s = str(s).strip()
    s_lower = s.lower()
    compact = "".join(ch for ch in s_lower if ch.isalnum())
    return s_lower, compact

def normalize_type(val):
    if val is None:
        return "OTHER"
    s, compact = _clean_key(val)
    if compact in _TYPE_MAP:
        return _TYPE_MAP[compact]
    if s in _TYPE_MAP:
        return _TYPE_MAP[s]
    if "non" in compact and ("res" in compact or "residential" in s):
        return "NON RES"
    if "res" in compact or "residential" in s:
        return "RES"
    return "OTHER"

def normalize_level(val):
    if val is None:
        return None
    s, compact = _clean_key(val)
    low = s
    for code, keywords in _LEVEL_RULES:
        for kw in keywords:
            if kw in low or kw.replace(" ", "") in compact:
                if code == "BLOCK_DISTRICT":
                    if ("block" in low or "block" in compact) and ("district" in low or "district" in compact):
                        return code
                    continue
                return code
    if compact in ("shg","clf","village"):
        return compact.upper()
    return None

def normalize_approval(val):
    if val is None:
        return None
    s, compact = _clean_key(val)
    for k, code in _APPROVAL_MAP.items():
        if k in s:
            return code
    if "sanction" in s:
        return "SANCTIONED"
    if "pend" in s:
        return "PENDING"
    if "deny" in s or "reject" in s:
        return "DENIED"
    return None

def try_parse_int(val):
    if val is None or val == "":
        return None
    try:
        return int(float(str(val)))
    except Exception:
        return None

def try_parse_date(val):
    if val is None or val == "":
        return None
    if hasattr(val, "year") and hasattr(val, "month"):
        return val.date() if hasattr(val, "date") else val
    s = str(val).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None

# --- management command ----------------------------
class Command(BaseCommand):
    help = "Import TrainingPlan rows from CSV or XLSX. Normalizes type_of_training case-insensitively. Leaves theme_expert NULL."

    def add_arguments(self, parser):
        parser.add_argument("file", help="Path to CSV or XLSX file")
        parser.add_argument("--format", choices=("csv", "xlsx"), default=None, help="force format")
        parser.add_argument("--update", action="store_true", help="update existing rows (match by training_name). If not set, existing names are skipped.")
        parser.add_argument("--dry-run", action="store_true", help="validate and show mapping, but do not write to DB")
        parser.add_argument("--preview", type=int, default=0, help="print the first N processed rows for inspection")

    def handle(self, *args, **options):
        p = Path(options["file"])
        if not p.exists():
            self.stderr.write(f"File does not exist: {p}")
            return 1

        fmt = options["format"]
        if fmt is None:
            if p.suffix.lower() == ".csv":
                fmt = "csv"
            elif p.suffix.lower() in (".xls", ".xlsx"):
                fmt = "xlsx"
            else:
                self.stderr.write("Couldn't infer file format. Use --format csv|xlsx")
                return 1

        do_update = options["update"]
        dry_run = options["dry_run"]
        preview_n = options["preview"]

        EXPECTED = {
            "training_name": ["training_name", "module", "module(training name)", "module (training name)"],
            "theme": ["theme"],
            "type_of_training": ["type", "type_of_training", "type of training"],
            "level_of_training": ["level", "level_of_training", "level of training"],
            "no_of_days": ["days", "no_of_days", "no of days"],
            "approval_status": ["approval", "approval_status", "approval of training plan"],
        }

        def header_lookup_map(fieldnames):
            mapping = {}
            lower_headers = {h.strip().lower(): h for h in fieldnames if h}
            for key, possibilities in EXPECTED.items():
                found = None
                for cand in possibilities:
                    if cand.strip().lower() in lower_headers:
                        found = lower_headers[cand.strip().lower()]
                        break
                mapping[key] = found
            return mapping

        rows = []
        map_hdr = {}
        if fmt == "csv":
            with open(p, newline='', encoding="utf-8") as fh:
                reader = csv.DictReader(fh)
                headers = reader.fieldnames
                map_hdr = header_lookup_map(headers)
                self.stdout.write(f"Detected headers: {headers}")
                for row in reader:
                    rows.append(row)
        else:
            wb = load_workbook(filename=str(p), read_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                self.stderr.write("Empty XLSX file")
                return 1
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            map_hdr = header_lookup_map(headers)
            self.stdout.write(f"Detected headers: {headers}")

            def gen():
                for r in it:
                    d = {}
                    for i, val in enumerate(r):
                        key = headers[i] if i < len(headers) else f"col{i}"
                        d[str(key)] = val
                    yield d
            rows = list(gen())

        total = len(rows)
        self.stdout.write(f"Found {total} rows. Starting import (update={do_update}, dry_run={dry_run})")

        processed = 0
        created = 0
        updated = 0
        skipped = 0
        skipped_names = []
        previewed = 0

        # main loop
        for r in rows:
            processed += 1

            def get_val(expected_key):
                hdr = map_hdr.get(expected_key)
                if not hdr:
                    return None
                return r.get(hdr)

            training_name = get_val("training_name")
            if training_name is None:
                self.stderr.write(f"Row {processed}: missing training_name; skipping")
                skipped += 1
                continue
            training_name = str(training_name).strip()

            theme = get_val("theme") or None
            if theme is not None:
                theme = str(theme).strip() or None

            raw_type = get_val("type_of_training")
            normalized_type = normalize_type(raw_type)

            raw_level = get_val("level_of_training")
            normalized_level = normalize_level(raw_level)

            no_of_days_raw = get_val("no_of_days")
            no_of_days = try_parse_int(no_of_days_raw)

            approval_raw = get_val("approval_status")
            approval = normalize_approval(approval_raw)

            if preview_n and previewed < preview_n:
                self.stdout.write(f"[preview row {processed}] name={training_name!r}, raw_type={raw_type!r} -> {normalized_type}, level_raw={raw_level!r} -> {normalized_level}, days={no_of_days}, approval={approval}, theme={theme}")
                previewed += 1

            # If dry_run, we still check DB existence so dry-run output matches real run
            exists = TrainingPlan.objects.filter(training_name=training_name).exists()

            if dry_run:
                if exists:
                    skipped += 1
                    skipped_names.append(training_name)
                else:
                    created += 1
                continue

            defaults = {
                "theme": theme,
                "type_of_training": normalized_type,
                "level_of_training": normalized_level,
                "no_of_days": no_of_days,
                "approval_status": approval,
            }

            try:
                if do_update:
                    obj, created_flag = TrainingPlan.objects.update_or_create(
                        training_name=training_name,
                        defaults=defaults
                    )
                    if created_flag:
                        created += 1
                    else:
                        updated += 1
                else:
                    if exists:
                        skipped += 1
                        skipped_names.append(training_name)
                    else:
                        TrainingPlan.objects.create(training_name=training_name, **{k: v for k, v in defaults.items() if v is not None})
                        created += 1
            except IntegrityError as e:
                self.stderr.write(f"Row {processed} DB error: {e}; skipping")
                skipped += 1
            except Exception as e:
                self.stderr.write(f"Row {processed} unexpected error: {e}; skipping")
                skipped += 1

        if skipped_names:
            self.stdout.write("Skipped (already existed):")
            for name in skipped_names:
                self.stdout.write(" - " + name)

        self.stdout.write(self.style.SUCCESS(
            f"Done. processed={processed}, created={created}, updated={updated}, skipped={skipped}"
        ))
        return 0
