# bmmu/management/commands/import_beneficiaries.py
import csv
import sys
from datetime import datetime, date
from pathlib import Path
from django.core.management.base import BaseCommand
from django.db import transaction, IntegrityError, connection
from openpyxl import load_workbook

from bmmu.models import Beneficiary

# === Field map: header in your file -> model field name ===
FIELD_MAP = {
    "state": "state",
    "district": "district",
    "block": "block",
    "gram_panchayat": "gram_panchayat",
    "village": "village",
    "shg_code": "shg_code",
    "shg_name": "shg_name",
    "date_of_formation": "date_of_formation",
    "member_code": "member_code",
    "member_name": "member_name",
    "date_of_birth": "date_of_birth",
    "date_of_joining_shg": "date_of_joining_shg",
    "designation_in_shg_vo_clf": "designation_in_shg_vo_clf",
    "social_category": "social_category",
    "pvtg_category": "pvtg_category",
    "religion": "religion",
    "gender": "gender",
    "education": "education",
    "marital_status": "marital_status",
    "insurance_status": "insurance_status",
    "disability": "disability",
    "disability_type": "disability_type",
    "is_head_of_family": "is_head_of_family",
    "parent_or_spouse_name": "parent_or_spouse_name",
    "relation": "relation",
    "account_number": "account_number",
    "ifsc": "ifsc",
    "branch_name": "branch_name",
    "bank_name": "bank_name",
    "account_opening_date": "account_opening_date",
    "account_type": "account_type",
    "mobile_no": "mobile_no",
    "aadhaar_no": "aadhaar_no",
    "aadhar_kyc": "aadhar_kyc",
    "ekyc_status": "ekyc_status",
    "cadres_role": "cadres_role",
    "primary_livelihood": "primary_livelihood",
    "secondary_livelihood": "secondary_livelihood",
    "tertiary_livelihood": "tertiary_livelihood",
    "nrega_no": "nrega_no",
    "pmay_id": "pmay_id",
    "secc_tin": "secc_tin",
    "nrlm_id": "nrlm_id",
    "state_id": "state_id",
    "ebk_id": "ebk_id",
    "ebk_name": "ebk_name",
    "ebk_mobile_no": "ebk_mobile_no",
    "approval_status": "approval_status",
    "date_of_approval": "date_of_approval",
    "benef_status": "benef_status",
    "inactive_date": "inactive_date",
    "inactive_reason": "inactive_reason",
    "member_type": "member_type",
}

# date fields that should be converted to date objects if possible
DATE_FIELDS = {"date_of_formation", "date_of_birth", "date_of_joining_shg", "account_opening_date", "date_of_approval", "inactive_date"}

# common date formats to try
DATE_FORMATS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y", "%d %b %Y", "%d %B %Y")

DEFAULT_CHUNK = 2000

def try_parse_date(val):
    if val is None or val == "":
        return None
    # if openpyxl already returned a date/datetime object
    if hasattr(val, "year") and hasattr(val, "month"):
        return val.date() if hasattr(val, "date") else val
    s = str(val).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    # final fallback: try YYYYMMDD or numeric floats from Excel
    try:
        if s.isdigit() and len(s) == 8:
            return datetime.strptime(s, "%Y%m%d").date()
    except Exception:
        pass
    # give up — return original string (DB will accept if model field is CharField, otherwise None)
    return None

def normalize_mobile(s):
    if s is None:
        return ""
    s = str(s)
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits or ""

class Command(BaseCommand):
    help = "Import beneficiaries from CSV or XLSX using chunked bulk_create. Safe for large files."

    def add_arguments(self, parser):
        parser.add_argument("file", help="Path to CSV or XLSX file")
        parser.add_argument("--format", choices=("csv", "xlsx"), default=None)
        parser.add_argument("--chunk", type=int, default=DEFAULT_CHUNK, help="chunk size for bulk_create")
        parser.add_argument("--ignore-duplicates", action="store_true", help="use ignore_conflicts=True to skip unique constraint conflicts")
        parser.add_argument("--fast-sqlite", action="store_true", help="(sqlite only) set PRAGMA journal_mode=WAL and synchronous=NORMAL for faster writes")

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.exists():
            self.stderr.write(f"File not found: {path}")
            return 1

        fmt = options["format"]
        if fmt is None:
            if path.suffix.lower() == ".csv":
                fmt = "csv"
            elif path.suffix.lower() in (".xls", ".xlsx"):
                fmt = "xlsx"
            else:
                self.stderr.write("Unable to infer format. Use --format csv|xlsx")
                return 1

        chunk = options["chunk"]
        ignore_conflicts = options["ignore_duplicates"]
        fast_sqlite = options["fast_sqlite"]

        # optional sqlite pragmas
        if fast_sqlite and "sqlite" in connection.settings_dict["ENGINE"]:
            try:
                with connection.cursor() as cur:
                    cur.execute("PRAGMA journal_mode = WAL;")
                    cur.execute("PRAGMA synchronous = NORMAL;")
                self.stdout.write("SQLite pragmas set (WAL / synchronous=NORMAL).")
            except Exception as e:
                self.stdout.write(f"Could not set SQLite pragmas: {e}")

        # prepare error log CSV
        errors_path = Path.cwd() / "import_errors.csv"
        err_f = open(errors_path, "w", newline='', encoding="utf-8")
        err_writer = None

        def process_row_to_instance(row):
            kwargs = {}
            for header, model_field in FIELD_MAP.items():
                # case-insensitive lookup
                val = None
                if header in row:
                    val = row.get(header)
                else:
                    # try lower-match
                    for k in row:
                        if k and k.strip().lower() == header.lower():
                            val = row.get(k)
                            break
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    continue
                # dates
                if header in DATE_FIELDS:
                    parsed = try_parse_date(val)
                    if parsed:
                        kwargs[model_field] = parsed
                    # else leave None (or keep model default)
                elif model_field == "mobile_no" or model_field == "ebk_mobile_no":
                    kwargs[model_field] = normalize_mobile(val)
                else:
                    kwargs[model_field] = str(val).strip()
            return Beneficiary(**kwargs)

        total = 0
        created = 0
        skipped = 0
        buffer = []

        try:
            if fmt == "csv":
                f = open(path, newline='', encoding='utf-8')
                reader = csv.DictReader(f)
                header_names = reader.fieldnames
                iterator = reader
            else:  # xlsx
                wb = load_workbook(filename=str(path), read_only=True)
                ws = wb.active
                it = ws.iter_rows(values_only=True)
                try:
                    header_row = next(it)
                except StopIteration:
                    self.stderr.write("Empty xlsx")
                    return 1
                header_names = [str(h).strip() if h is not None else "" for h in header_row]
                def gen():
                    for row in it:
                        d = {}
                        for i, cell in enumerate(row):
                            key = header_names[i] if i < len(header_names) else f"col{i}"
                            d[str(key)] = cell
                        yield d
                iterator = gen()

            self.stdout.write(f"Detected headers: {header_names}")
            self.stdout.write(f"Starting import: chunk={chunk}, ignore_conflicts={ignore_conflicts}")

            for row in iterator:
                total += 1
                try:
                    inst = process_row_to_instance(row)
                    buffer.append(inst)
                except Exception as e:
                    # log the raw row + error
                    if err_writer is None:
                        err_writer = csv.DictWriter(err_f, fieldnames=list(row.keys()) + ["error"])
                        err_writer.writeheader()
                    row_copy = {k: (v if not isinstance(v, bytes) else v.decode(errors="ignore")) for k,v in row.items()}
                    row_copy["error"] = f"process_error:{e}"
                    err_writer.writerow(row_copy)
                    skipped += 1
                    continue

                if len(buffer) >= chunk:
                    with transaction.atomic():
                        if ignore_conflicts:
                            Beneficiary.objects.bulk_create(buffer, ignore_conflicts=True)
                        else:
                            Beneficiary.objects.bulk_create(buffer)
                    created += len(buffer)
                    self.stdout.write(f"[{total}] Inserted so far: {created}")
                    buffer = []

            # flush remainder
            if buffer:
                with transaction.atomic():
                    if ignore_conflicts:
                        Beneficiary.objects.bulk_create(buffer, ignore_conflicts=True)
                    else:
                        Beneficiary.objects.bulk_create(buffer)
                created += len(buffer)
                self.stdout.write(f"Final flush inserted {len(buffer)} rows.")

        except IntegrityError as e:
            self.stderr.write("IntegrityError during import: " + str(e))
            self.stderr.write("Try --ignore-duplicates or inspect import_errors.csv")
            return 1
        except Exception as e:
            self.stderr.write("Fatal error: " + str(e))
            raise
        finally:
            err_f.close()

        self.stdout.write(self.style.SUCCESS(f"Done. Total read: {total}, created (attempted inserts): {created}, skipped (processing errors): {skipped}"))
        self.stdout.write(f"Any processing errors logged to: {errors_path}")
        return 0
